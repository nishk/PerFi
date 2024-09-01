import argparse
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import yaml
from gspread_formatting import set_column_width, format_cell_range, CellFormat, Color


# Load the configuration from the input.yaml file
def load_config():
    with open("input.yaml", 'r') as stream:
        try:
            config = yaml.safe_load(stream)
            return config
        except yaml.YAMLError as exc:
            print(f"Error loading YAML configuration: {exc}")
            exit(1)

# Define contribution limits for 2024 and 2025
contribution_limits = {
    2024: {
        "HSA_INDIVIDUAL": 4150,
        "HSA_FAMILY": 8300,
        "K401": 23000
    },
    2025: {
        "HSA_INDIVIDUAL": 4300,  # Example values for 2025
        "HSA_FAMILY": 8600,
        "K401": 24000
    }
}

# Function to calculate remaining contribution or excess
def calculate_contribution_status(contribution_to_date, limit):
    remaining = limit - contribution_to_date
    if remaining >= 0:
        return f"Remaining Contribution: ${remaining:.2f}", round(remaining, 2)
    else:
        return f"Exceeded Contribution: ${abs(remaining):.2f}", round(-abs(remaining), 2)

# Function to output the data to an Excel file
def save_contribution_to_excel(hsa_contributed, hsa_status, hsa_amount, k401_contributed, k401_status, k401_amount, is_family, file_path, year, HSA_LIMIT, K401_LIMIT):
    contribution_type_hsa = f"HSA {'Family' if is_family else 'Individual'}"
    contribution_type_k401 = "401(k) Individual"

    data = {
        'Contribution Type': [contribution_type_hsa, contribution_type_k401],
        'Contributed To Date ($)': [hsa_contributed, k401_contributed],
        'Status': [hsa_status.split(":")[0], k401_status.split(":")[0]],
        'Amount ($)': [abs(hsa_amount), abs(k401_amount)],
        'Contribution Limit ($)': [HSA_LIMIT, K401_LIMIT]
    }
    
    df = pd.DataFrame(data)
    
    # Set workbook and sheet names
    workbook_name = os.path.join(file_path, "contribution_summary.xlsx")
    date_str = datetime.today().strftime('%Y-%m-%d')
    sheet_name = f"{year}_Summary_{date_str}"

    # Check if the workbook exists
    if os.path.exists(workbook_name):
        # Load the existing workbook
        workbook = load_workbook(workbook_name)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]  # Load the existing sheet
            workbook.remove(worksheet)  # Remove existing sheet to overwrite
            worksheet = workbook.create_sheet(sheet_name)  # Recreate sheet
        else:
            worksheet = workbook.create_sheet(sheet_name)
    else:
        # Create a new workbook and sheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    # Write the headers with bold font size 16
    for c_idx, header in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=c_idx)
        cell.value = header
        cell.font = Font(size=16, bold=True)

    # Write the DataFrame to the sheet with font size 14
    for r_idx, row in enumerate(df.values, 2):  # Start at row 2 to account for header
        for c_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(size=14)

    # Highlight cells in the "Status" column if the contribution is exceeded
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    for row in range(2, len(df) + 2):  # Start from the second row (first row is the header)
        status_cell = worksheet[f'C{row}']  # "Status" column is C
        if "Exceeded" in status_cell.value:
            status_cell.fill = red_fill

    # Adjust column width to fit the content, including considering font size
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column letter
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2  # Add some padding and adjust
        worksheet.column_dimensions[column].width = adjusted_width

    # Set the default zoom level
    worksheet.sheet_view.zoomScale = 120  # Example: Set zoom to 120%

    # Save the workbook
    workbook.save(workbook_name)
    print(f"Contribution data saved to {workbook_name}, sheet {sheet_name}")
    
    return workbook_name

# Function to convert column number to letter (A, B, C...)
def get_column_letter(col_num):
    string = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        string = chr(65 + remainder) + string
    return string

# Function to auto-resize columns based on content length
def auto_resize_columns(worksheet):
    for col in range(1, worksheet.col_count + 1):
        max_len = 0
        col_values = worksheet.col_values(col)
        for value in col_values:
            if len(str(value)) > max_len:
                max_len = len(str(value))
        # Convert column number to letter and set the column width
        col_letter = get_column_letter(col)
        pixel_width = max(80, min(400, max_len * 10))
        set_column_width(worksheet, col_letter, pixel_width)

# Function to update the Google Sheet
def update_google_sheet(hsa_contributed, hsa_status, hsa_amount, k401_contributed, k401_status, k401_amount, is_family, google_sheet_url, year, HSA_LIMIT, K401_LIMIT, credentials_file):
    # Authenticate and open the Google Sheet
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_url(google_sheet_url)

    date_str = datetime.today().strftime('%Y-%m-%d')
    sheet_name = f"{year}_Summary_{date_str}"

    try:
        worksheet = sheet.worksheet(sheet_name)
        sheet.del_worksheet(worksheet)
        worksheet = sheet.add_worksheet(title=sheet_name, rows="100", cols="20")
    except gspread.exceptions.WorksheetNotFound:
        worksheet = sheet.add_worksheet(title=sheet_name, rows="100", cols="20")

    # Prepare the data
    headers = ["Contribution Type", "Contributed To Date ($)", "Status", "Amount ($)", "Contribution Limit ($)"]
    data = [
        [f"HSA {'Family' if is_family else 'Individual'}", hsa_contributed, hsa_status.split(":")[0], abs(hsa_amount), HSA_LIMIT],
        ["401(k) Individual", k401_contributed, k401_status.split(":")[0], abs(k401_amount), K401_LIMIT]
    ]

    # Update the Google Sheet
    worksheet.append_row(headers)
    worksheet.append_rows(data)

    # Apply red fill for exceeded contributions
    red_fill_format = CellFormat(
        backgroundColor=Color(1, 0.8, 0.8)
    )
    if "Exceeded" in hsa_status:
        format_cell_range(worksheet, 'C2', red_fill_format)
    if "Exceeded" in k401_status:
        format_cell_range(worksheet, 'C3', red_fill_format)

    # Manually auto-resize the columns based on content
    auto_resize_columns(worksheet)

    print(f"Google Sheet updated: {google_sheet_url}, sheet {sheet_name}")
    # Authenticate and open the Google Sheet
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_url(google_sheet_url)

    date_str = datetime.today().strftime('%Y-%m-%d')
    sheet_name = f"{year}_Summary_{date_str}"

    try:
        worksheet = sheet.worksheet(sheet_name)
        sheet.del_worksheet(worksheet)
        worksheet = sheet.add_worksheet(title=sheet_name, rows="100", cols="20")
    except gspread.exceptions.WorksheetNotFound:
        worksheet = sheet.add_worksheet(title=sheet_name, rows="100", cols="20")

    # Prepare the data
    headers = ["Contribution Type", "Contributed To Date ($)", "Status", "Amount ($)", "Contribution Limit ($)"]
    data = [
        [f"HSA {'Family' if is_family else 'Individual'}", hsa_contributed, hsa_status.split(":")[0], abs(hsa_amount), HSA_LIMIT],
        ["401(k) Individual", k401_contributed, k401_status.split(":")[0], abs(k401_amount), K401_LIMIT]
    ]

    # Update the Google Sheet
    worksheet.append_row(headers)
    worksheet.append_rows(data)

    # Apply red fill for exceeded contributions
    red_fill_format = CellFormat(
        backgroundColor=Color(1, 0.8, 0.8)
    )
    if "Exceeded" in hsa_status:
        format_cell_range(worksheet, 'C2', red_fill_format)
    if "Exceeded" in k401_status:
        format_cell_range(worksheet, 'C3', red_fill_format)

    # Manually auto-resize the columns based on content
    auto_resize_columns(worksheet)

    print(f"Google Sheet updated: {google_sheet_url}, sheet {sheet_name}")

# Function to convert column number to letter (A, B, C...)
def get_column_letter(col_num):
    string = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        string = chr(65 + remainder) + string
    return string

# Main function
def main():
    # Load configuration from input.yaml
    config = load_config()

    # Setup argument parser
    parser = argparse.ArgumentParser(description="Calculate remaining HSA and 401(k) contributions for 2024 or 2025.")
    parser.add_argument('--year', type=int, choices=[2024, 2025], required=True, help="Year for which you want to calculate the contributions (2024 or 2025).")
    parser.add_argument('--hsa', type=float, metavar='HSA contribution amount', required=True, help="Amount contributed to HSA so far in the selected year.")
    parser.add_argument('--k401', type=float, metavar='401K contribution limit', required=True, help="Amount contributed to 401(k) so far in the selected year.")
    parser.add_argument('--family', action='store_true', help="Flag indicating if HSA is for family (if not set, assumed individual).")

    args = parser.parse_args()

    # Ensure at least one of file_path or google_sheet_url is specified
    if not config.get('file_path') and not config.get('google_sheet_url'):
        print("Error: At least one of 'file_path' or 'google_sheet_url' must be specified in input.yaml.")
        return

    # Get the contribution limits for the selected year
    limits = contribution_limits[args.year]
    HSA_LIMIT = limits["HSA_FAMILY"] if args.family else limits["HSA_INDIVIDUAL"]
    K401_LIMIT = limits["K401"]

    # Calculate HSA contribution status
    hsa_status, hsa_amount = calculate_contribution_status(args.hsa, HSA_LIMIT)

    # Calculate 401(k) contribution status
    k401_status, k401_amount = calculate_contribution_status(args.k401, K401_LIMIT)

    # Output results to terminal
    hsa_type = "HSA Family" if args.family else "HSA Individual"
    print(f"{hsa_type}: {hsa_status}")
    print(f"401(k): {k401_status}")

    # Save to Excel if file_path is provided
    if config.get('file_path'):
        file_name = save_contribution_to_excel(args.hsa, hsa_status, hsa_amount, args.k401, k401_status, k401_amount, args.family, config['file_path'], args.year, HSA_LIMIT, K401_LIMIT)

    # Update Google Sheet if google_sheet_url is provided
    if config.get('google_sheet_url'):
        update_google_sheet(
            args.hsa, hsa_status, hsa_amount,
            args.k401, k401_status, k401_amount,
            args.family,
            config['google_sheet_url'],
            args.year,
            HSA_LIMIT,
            K401_LIMIT,
            config['credentials_file']
        )

if __name__ == "__main__":
    main()